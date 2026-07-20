"""
parser.py - Core PDF parsing logic for Bank Indonesia (Mandiri) credit card statements.

Extracts transaction data from PDF statements, handling:
- Multi-line descriptions (e.g. currency conversion info)
- Cross-page table continuations
- TAGIHAN BULAN LALU and SUB-TOTAL rows
- Various payment types (CR entries)
"""

import os
import re
import pdfplumber
from dataclasses import dataclass, field
from typing import Optional


@dataclass
class Transaction:
    """Represents a single transaction row."""
    tanggal_transaksi: str  # DD/MM/YY
    tanggal_pembukuan: str  # DD/MM/YY
    keterangan: str
    jumlah: float
    is_credit: bool = False  # True if amount has CR suffix


@dataclass
class CardholderRecord:
    """Represents one cardholder's complete statement."""
    nama: str
    nomor_kartu: str
    limit_kartu_kredit: Optional[float] = None
    tagihan_bulan_lalu: Optional[float] = None
    tagihan_bulan_lalu_cr: bool = False
    sub_total: Optional[float] = None
    sub_total_cr: bool = False
    transactions: list = field(default_factory=list)


# ── Regex Patterns ──────────────────────────────────────────────────────────

# Transaction line: DD/MM/YY DD/MM/YY description amount [CR]
RE_TRANSACTION = re.compile(
    r'^(\d{2}/\d{2}/\d{2})\s+(\d{2}/\d{2}/\d{2})\s+'  # two dates
    r'(.+?)\s+'                                          # description (non-greedy)
    r'([\d,]+(?:\.\d+)?)\s*'                             # amount with commas
    r'(CR)?\s*$'                                         # optional CR
)

# Alternative: amount might be glued to CR without space (e.g. "195,000CR")
RE_TRANSACTION_ALT = re.compile(
    r'^(\d{2}/\d{2}/\d{2})\s+(\d{2}/\d{2}/\d{2})\s+'
    r'(.+?)\s+'
    r'([\d,]+(?:\.\d+)?)(CR)?\s*$'
)

# TAGIHAN BULAN LALU line
RE_TAGIHAN = re.compile(
    r'^\s*TAGIHAN BULAN LALU\s+([\d,]+(?:\.\d+)?)\s*(CR)?\s*$'
)

# SUB-TOTAL line
RE_SUBTOTAL = re.compile(
    r'^\s*SUB-TOTAL\s+([\d,]+(?:\.\d+)?)\s*(CR)?\s*$'
)

# Card number pattern: XXXX-XXXX-XXXX-XXXX
RE_CARD_NUMBER = re.compile(r'^(\d{4}-\d{4}-\d{4}-\d{4})\s*$')

# Credit limit line: e.g. "50,000,000 50,010,000 0 LANCAR" or "50,000,000 236,724 CR 0 LANCAR"
RE_CREDIT_LIMIT = re.compile(r'^([\d,]+)\s+[\d,]+\s*(?:CR)?\s+\d+\s+LANCAR')

# Date pattern at start of line (to detect transaction start)
RE_DATE_START = re.compile(r'^\d{2}/\d{2}/\d{2}\s')

# Lines to skip
RE_SKIP_PATTERNS = [
    re.compile(r'^\.\.\.\.\.\.\.\s*Transaksi di lanjutkan'),
    re.compile(r'^1\s+(?:Apabila|Selain|Sesuai|Jaga)'),
    re.compile(r'^New Livin'),
    re.compile(r'^kartu kredit'),  # continuation of disclaimer
    re.compile(r'^kredit, kode OTP'),
    re.compile(r'^akumulasi'),
    re.compile(r'^Pagu Kredit'),
    re.compile(r'^Credit Limit'),
    re.compile(r'^\d+[,.]?\d*[,.]?\d*\s+\d+[,.]?\d*[,.]?\d*\s+\d+\s+LANCAR'),
    re.compile(r'^1\.750%'),
    re.compile(r'^21\.00%'),
    re.compile(r'^atau$'),
    re.compile(r'^Gunakanlah'),
    re.compile(r'^melalui mandiri'),
    re.compile(r'^14000'),
    re.compile(r'^terdekat'),
    re.compile(r'^yang ditujukan'),
    re.compile(r'^Tanggal Transaksi'),
    re.compile(r'^Transaction Date'),
    re.compile(r'^Tanggal Pembukaan'),  # continuation page header variant
    re.compile(r'^Nomor Kartu'),
    re.compile(r'^Account Number'),
    re.compile(r'^dsdsfd'),
    re.compile(r'^BAPAK\s'),
    re.compile(r'^IBU\s'),
    re.compile(r'^BANK INDONESIA'),
    re.compile(r'^GD\s'),
    re.compile(r'^JL\s'),
    re.compile(r'^KEL\.\s'),
    re.compile(r'^GAMBIR'),
    re.compile(r'^JAKARTA\s'),
    re.compile(r'^\d{3}\s+\d{2,3}\s*$'),  # "000 000" or "000 00"
    re.compile(r'^Tagihan Baru'),
    re.compile(r'^New balance'),
    re.compile(r'^Pembayaran Minimum'),
    re.compile(r'^Minimum payment'),
    re.compile(r'^Tanggal Cetak'),
    re.compile(r'^Statement Date'),
    re.compile(r'^Tanggal Jatuh'),
    re.compile(r'^Payment Due'),
    re.compile(r'^Sisa Pagu'),
    re.compile(r'^Sisa Tagihan'),
    re.compile(r'^Remaining'),
    re.compile(r'^Kualitas Kredit'),
    re.compile(r'^Loan Performance'),
    re.compile(r'^Bunga Pembelanjaan'),
    re.compile(r'^Interest Rate'),
]


def parse_amount(amount_str: str) -> float:
    """Convert comma-formatted amount string to float."""
    return float(amount_str.replace(',', ''))


def is_skip_line(line: str) -> bool:
    """Check if a line should be skipped (disclaimers, headers, etc.)."""
    stripped = line.strip()
    if not stripped:
        return True
    for pattern in RE_SKIP_PATTERNS:
        if pattern.search(stripped):
            return True
    return False


def is_continuation_page(text: str) -> bool:
    """
    Determine if a page is a continuation page (no cardholder header).
    Continuation pages only have the transaction table header + data.
    """
    # Main pages have the billing summary with "Nomor Kartu" or card number info
    # and "Gunakanlah kemudahan"
    has_billing_header = 'Gunakanlah kemudahan' in text
    return not has_billing_header


def extract_cardholder_info(lines: list[str]) -> tuple[str, str]:
    """
    Extract cardholder name and card number from text lines.
    
    In the statement, the section between the transaction table header 
    and the first transaction contains:
    - Cardholder name (e.g., "AGUNG HARTONO")
    - TAGIHAN BULAN LALU line
    - Card number (e.g., "4895-9403-0012-6163")
    """
    name = ""
    card_number = ""
    
    # Find the transaction table header position
    header_idx = -1
    for i, line in enumerate(lines):
        if 'Tanggal Transaksi' in line and 'Keterangan' in line:
            header_idx = i
            break
    
    if header_idx == -1:
        return name, card_number
    
    # After header, skip "Transaction Date..." line, then next non-empty line is the name
    # The name appears right after the header row
    search_start = header_idx + 1
    for i in range(search_start, min(search_start + 5, len(lines))):
        line = lines[i].strip()
        # Skip the English translation header
        if line.startswith('Transaction Date'):
            continue
        if not line:
            continue
        # This should be the cardholder name
        # It's NOT a date line, NOT a TAGIHAN line, NOT a card number
        if not RE_DATE_START.match(line) and not RE_TAGIHAN.match(line) and not RE_CARD_NUMBER.match(line):
            # Could be name like "AGUNG HARTONO" or ". ABUDAUD GAMGULU"
            name = line.strip('. ')
            break
    
    # Find card number
    for line in lines:
        m = RE_CARD_NUMBER.match(line.strip())
        if m:
            card_number = m.group(1)
            break
    
    # If we found card number in the billing summary area, we need the one in the transaction section
    # Actually, both are the same card number, so the first match is fine
    
    return name, card_number


def parse_transaction_section(lines: list[str], start_idx: int) -> tuple[list[Transaction], Optional[float], bool, Optional[float], bool]:
    """
    Parse the transaction section of a page starting from start_idx.
    
    Returns:
        (transactions, tagihan_amount, tagihan_cr, subtotal_amount, subtotal_cr)
    """
    transactions = []
    tagihan_amount = None
    tagihan_cr = False
    subtotal_amount = None
    subtotal_cr = False
    
    current_tx = None
    
    for i in range(start_idx, len(lines)):
        line = lines[i].strip()
        
        if not line:
            continue
        
        # Check for SUB-TOTAL
        m_sub = RE_SUBTOTAL.match(line)
        if m_sub:
            subtotal_amount = parse_amount(m_sub.group(1))
            subtotal_cr = m_sub.group(2) == 'CR'
            # Finalize current transaction if any
            if current_tx:
                transactions.append(current_tx)
                current_tx = None
            break  # SUB-TOTAL is always the last line of transactions
        
        # Check for TAGIHAN BULAN LALU
        m_tag = RE_TAGIHAN.match(line)
        if m_tag:
            tagihan_amount = parse_amount(m_tag.group(1))
            tagihan_cr = m_tag.group(2) == 'CR'
            continue
        
        # Skip card number line
        if RE_CARD_NUMBER.match(line):
            continue
        
        # Skip irrelevant lines
        if is_skip_line(line):
            # If we hit a disclaimer/footer area, finalize
            if current_tx and (line.startswith('1 ') or 'Pagu Kredit' in line):
                transactions.append(current_tx)
                current_tx = None
            continue
        
        # Try to match a transaction line
        m_tx = RE_TRANSACTION.match(line)
        if not m_tx:
            m_tx = RE_TRANSACTION_ALT.match(line)
        
        if m_tx:
            # Finalize previous transaction
            if current_tx:
                transactions.append(current_tx)
            
            current_tx = Transaction(
                tanggal_transaksi=m_tx.group(1),
                tanggal_pembukuan=m_tx.group(2),
                keterangan=m_tx.group(3).strip(),
                jumlah=parse_amount(m_tx.group(4)),
                is_credit=(m_tx.group(5) == 'CR') if m_tx.group(5) else False
            )
            continue
        
        # If we're here, this is a continuation line for the current transaction description
        if current_tx:
            # Append to description
            current_tx.keterangan += '\n' + line
        # else: could be the cardholder name or other non-transaction text, skip
    
    # Finalize last transaction
    if current_tx:
        transactions.append(current_tx)
    
    return transactions, tagihan_amount, tagihan_cr, subtotal_amount, subtotal_cr


def find_transaction_start(lines: list[str]) -> int:
    """Find the line index where transaction data begins."""
    for i, line in enumerate(lines):
        stripped = line.strip()
        # Look for the English translation header line
        if stripped.startswith('Transaction Date') and 'Posting Date' in stripped:
            return i + 1  # transactions start after this
    return 0


def parse_pdf(filepath: str) -> list[CardholderRecord]:
    """
    Parse a Bank Indonesia (Mandiri) credit card statement PDF.
    
    Uses chunked processing to keep memory usage low:
    1. PyMuPDF splits the PDF into small temporary chunks
    2. Each chunk is processed with pdfplumber (which produces correct table layout)
    3. Memory is freed between chunks via gc.collect()
    
    Returns a list of CardholderRecord objects, one per cardholder.
    """
    import fitz  # PyMuPDF - lightweight, used only for splitting
    import tempfile
    import gc
    
    records = []
    current_record = None
    
    CHUNK_SIZE = 15  # process 15 pages at a time to keep memory low
    
    doc = fitz.open(filepath)
    total_pages = len(doc)
    
    for chunk_start in range(0, total_pages, CHUNK_SIZE):
        chunk_end = min(chunk_start + CHUNK_SIZE, total_pages)
        
        # Create a temporary PDF with just this chunk of pages
        chunk_doc = fitz.open()
        chunk_doc.insert_pdf(doc, from_page=chunk_start, to_page=chunk_end - 1)
        
        tmp_fd, tmp_path = tempfile.mkstemp(suffix='.pdf')
        os.close(tmp_fd)
        chunk_doc.save(tmp_path)
        chunk_doc.close()
        
        # Process this chunk with pdfplumber
        try:
            with pdfplumber.open(tmp_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if not text:
                        continue
                    
                    lines = text.split('\n')
                    
                    if is_continuation_page(text):
                        if current_record:
                            start_idx = find_transaction_start(lines)
                            txs, _, _, subtotal, subtotal_cr = parse_transaction_section(lines, start_idx)
                            current_record.transactions.extend(txs)
                            if subtotal is not None:
                                current_record.sub_total = subtotal
                                current_record.sub_total_cr = subtotal_cr
                    else:
                        if current_record:
                            records.append(current_record)
                        
                        name, card_number = extract_cardholder_info(lines)
                        
                        credit_limit = None
                        for line in lines:
                            m_limit = RE_CREDIT_LIMIT.match(line.strip())
                            if m_limit:
                                credit_limit = parse_amount(m_limit.group(1))
                                break
                        
                        current_record = CardholderRecord(
                            nama=name,
                            nomor_kartu=card_number,
                            limit_kartu_kredit=credit_limit
                        )
                        
                        start_idx = find_transaction_start(lines)
                        txs, tagihan, tagihan_cr, subtotal, subtotal_cr = parse_transaction_section(lines, start_idx)
                        
                        current_record.transactions = txs
                        current_record.tagihan_bulan_lalu = tagihan
                        current_record.tagihan_bulan_lalu_cr = tagihan_cr
                        if subtotal is not None:
                            current_record.sub_total = subtotal
                            current_record.sub_total_cr = subtotal_cr
        finally:
            # Clean up temp file and free memory
            try:
                os.remove(tmp_path)
            except OSError:
                pass
            gc.collect()
    
    doc.close()
    
    # Don't forget the last record
    if current_record:
        records.append(current_record)
    
    return records


def generate_excel(records: list[CardholderRecord], output_path: str) -> dict:
    """
    Generate an Excel file from parsed cardholder records.
    
    Creates a clean flat table suitable for pivot tables with columns:
    - Nama Pemegang Kartu
    - Nomor Kartu
    - Limit Kartu Kredit
    - Tanggal Transaksi
    - Tanggal Pembukuan
    - Keterangan
    - Jumlah (Rp)
    - Tipe (DB/CR)
    
    Returns stats dict with counts.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Transaksi"
    
    # ── Styles ──
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    section_font = Font(name='Calibri', bold=True, color='1F4E79', size=10)
    section_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    
    data_font = Font(name='Calibri', size=10)
    data_alignment = Alignment(vertical='center', wrap_text=True)
    amount_alignment = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='B4C6E7'),
        right=Side(style='thin', color='B4C6E7'),
        top=Side(style='thin', color='B4C6E7'),
        bottom=Side(style='thin', color='B4C6E7')
    )
    
    even_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
    credit_font = Font(name='Calibri', size=10, color='008000')  # Green for credits
    
    # ── Headers ──
    headers = [
        'Nama Pemegang Kartu',
        'Nomor Kartu',
        'Limit Kartu Kredit (Rp)',
        'Tanggal Transaksi',
        'Tanggal Pembukuan',
        'Keterangan',
        'Jumlah (Rp)',
        'Tipe'
    ]
    
    NUM_COLS = len(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # ── Column Widths ──
    col_widths = [25, 22, 22, 18, 18, 45, 18, 8]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # ── Data Rows ──
    row_num = 2
    total_transactions = 0
    total_cardholders = len(records)
    
    for record in records:
        limit_val = record.limit_kartu_kredit  # repeated on every row for pivot
        
        # TAGIHAN BULAN LALU row
        if record.tagihan_bulan_lalu is not None:
            tipe = 'CR' if record.tagihan_bulan_lalu_cr else 'DB'
            ws.cell(row=row_num, column=1, value=record.nama).font = section_font
            ws.cell(row=row_num, column=2, value=record.nomor_kartu).font = section_font
            c_limit = ws.cell(row=row_num, column=3, value=limit_val)
            c_limit.font = section_font
            c_limit.number_format = '#,##0'
            c_limit.alignment = amount_alignment
            ws.cell(row=row_num, column=4, value='').font = section_font
            ws.cell(row=row_num, column=5, value='').font = section_font
            ws.cell(row=row_num, column=6, value='TAGIHAN BULAN LALU').font = section_font
            ws.cell(row=row_num, column=7, value=record.tagihan_bulan_lalu).font = section_font
            ws.cell(row=row_num, column=7).number_format = '#,##0'
            ws.cell(row=row_num, column=7).alignment = amount_alignment
            ws.cell(row=row_num, column=8, value=tipe).font = section_font
            
            for col in range(1, NUM_COLS + 1):
                ws.cell(row=row_num, column=col).fill = section_fill
                ws.cell(row=row_num, column=col).border = thin_border
                ws.cell(row=row_num, column=col).alignment = data_alignment
            ws.cell(row=row_num, column=3).alignment = amount_alignment
            ws.cell(row=row_num, column=7).alignment = amount_alignment
            
            row_num += 1
        
        # Transaction rows
        for tx in record.transactions:
            is_even = (row_num % 2 == 0)
            tipe = 'CR' if tx.is_credit else 'DB'
            font = credit_font if tx.is_credit else data_font
            
            ws.cell(row=row_num, column=1, value=record.nama).font = font
            ws.cell(row=row_num, column=2, value=record.nomor_kartu).font = font
            c_limit = ws.cell(row=row_num, column=3, value=limit_val)
            c_limit.font = font
            c_limit.number_format = '#,##0'
            c_limit.alignment = amount_alignment
            ws.cell(row=row_num, column=4, value=tx.tanggal_transaksi).font = font
            ws.cell(row=row_num, column=5, value=tx.tanggal_pembukuan).font = font
            ws.cell(row=row_num, column=6, value=tx.keterangan).font = font
            ws.cell(row=row_num, column=7, value=tx.jumlah).font = font
            ws.cell(row=row_num, column=7).number_format = '#,##0'
            ws.cell(row=row_num, column=7).alignment = amount_alignment
            ws.cell(row=row_num, column=8, value=tipe).font = font
            
            for col in range(1, NUM_COLS + 1):
                ws.cell(row=row_num, column=col).border = thin_border
                ws.cell(row=row_num, column=col).alignment = data_alignment
                if is_even:
                    ws.cell(row=row_num, column=col).fill = even_fill
            ws.cell(row=row_num, column=3).alignment = amount_alignment
            ws.cell(row=row_num, column=7).alignment = amount_alignment
            
            row_num += 1
            total_transactions += 1
        
        # SUB-TOTAL row
        if record.sub_total is not None:
            tipe = 'CR' if record.sub_total_cr else 'DB'
            ws.cell(row=row_num, column=1, value=record.nama).font = section_font
            ws.cell(row=row_num, column=2, value=record.nomor_kartu).font = section_font
            c_limit = ws.cell(row=row_num, column=3, value=limit_val)
            c_limit.font = section_font
            c_limit.number_format = '#,##0'
            c_limit.alignment = amount_alignment
            ws.cell(row=row_num, column=4, value='').font = section_font
            ws.cell(row=row_num, column=5, value='').font = section_font
            ws.cell(row=row_num, column=6, value='SUB-TOTAL').font = section_font
            ws.cell(row=row_num, column=7, value=record.sub_total).font = section_font
            ws.cell(row=row_num, column=7).number_format = '#,##0'
            ws.cell(row=row_num, column=7).alignment = amount_alignment
            ws.cell(row=row_num, column=8, value=tipe).font = section_font
            
            for col in range(1, NUM_COLS + 1):
                ws.cell(row=row_num, column=col).fill = section_fill
                ws.cell(row=row_num, column=col).border = thin_border
                ws.cell(row=row_num, column=col).alignment = data_alignment
            ws.cell(row=row_num, column=3).alignment = amount_alignment
            ws.cell(row=row_num, column=7).alignment = amount_alignment
            
            row_num += 1
    
    # ── Freeze Panes & Auto-filter ──
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:H{row_num - 1}'
    
    # ── Save ──
    wb.save(output_path)
    
    return {
        'total_cardholders': total_cardholders,
        'total_transactions': total_transactions,
        'total_rows': row_num - 2,  # excluding header
        'output_path': output_path
    }


if __name__ == '__main__':
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python parser.py <pdf_path> [output_path]")
        sys.exit(1)
    
    pdf_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else 'output.xlsx'
    
    print(f"Parsing {pdf_path}...")
    records = parse_pdf(pdf_path)
    
    print(f"\nFound {len(records)} cardholders:")
    for r in records[:5]:
        print(f"  - {r.nama} ({r.nomor_kartu}): {len(r.transactions)} transactions")
    if len(records) > 5:
        print(f"  ... and {len(records) - 5} more")
    
    print(f"\nGenerating Excel: {output_path}")
    stats = generate_excel(records, output_path)
    print(f"Done! {stats['total_cardholders']} cardholders, {stats['total_transactions']} transactions, {stats['total_rows']} total rows")
